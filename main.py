import openpyxl
import pandas as pd
import csv

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


# Define the file path
file_path = "Directory.xlsx"

# Load the Excel spreadsheet into a pandas DataFrame
df = pd.read_excel(file_path)


# Define the function to apply on the DataFrame
def retention_allotment(enrollment):
    if isinstance(enrollment, (int, float)):
        if enrollment < 20000:
            return 6000.0
        elif enrollment >= 20000:
            return 2000.0
    else:
        return None


# Create a new column "Retention Allotment" and apply the function
df['Retention Allotment'] = df['District Enrollment as of Oct 2022'].apply(retention_allotment)


def get_classroom_teachers(district_number):

    with open("Staff Salary FTE Report_Statewide_Districts_2022-2023.csv", "r") as csvfile:
        reader = csv.reader(csvfile)

        header = False
        header_indexes = {}

        for row in reader:
            if header is False:
                header = True

                for i, header_label in enumerate(row):
                    header_indexes[header_label] = i
            else:
                dist_num_local = "'" + ("0" * (6-len(row[header_indexes["District"]]))) + row[header_indexes["District"]]

                if dist_num_local == district_number:

                    if row[header_indexes["Staff"]] == "TOTAL TEACHING STAFF":
                        return float(row[header_indexes["FTE Count"]])

    return None


df["Total Teaching Staff"] = df["District Number"].apply(get_classroom_teachers)

print(df["Total Teaching Staff"].head())

# Print the number of rows where the value in the "Total Teaching Staff" column is None
print("Rows with null Total Teaching Staff:", df['Total Teaching Staff'].isnull().sum())


# Create a separate dataframe with rows having null "Total Teaching Staff"
null_teaching_staff_df = df[df['Total Teaching Staff'].isnull()]

# Export the new dataframe to an Excel spreadsheet
null_teaching_staff_df.to_excel("Null_Total_Teaching_Staff.xlsx", index=False)

print("Rows with null Total Teaching Staff have been saved to 'Null_Total_Teaching_Staff.xlsx'")


def calculate_retention_allotment_cost(row):
    retention_allotment = row['Retention Allotment']
    total_teaching_staff = row['Total Teaching Staff']

    if retention_allotment is None or total_teaching_staff is None:
        return None
    else:
        return float(retention_allotment) * float(total_teaching_staff)


df["Retention Allotment Cost"] = df.apply(calculate_retention_allotment_cost, axis=1)

# Calculate the sum of the values in the "Retention Allotment Cost" column
sum_retention_allotment_cost = df['Retention Allotment Cost'].sum()

print("TOTAL COST:", sum_retention_allotment_cost)
# Calculate the sum of the values in the "Total Teaching Staff" column
sum_total_teaching_staff = df['Total Teaching Staff'].sum()

print("TOTAL FTE:", sum_total_teaching_staff)

print(" ")
# Calculate the average and print the result
average_retention_allotment_cost = sum_retention_allotment_cost / sum_total_teaching_staff
print("Average Retention Allotment Cost per Total Teaching Staff:", average_retention_allotment_cost)

print(" ")
print("Final total:", sum_retention_allotment_cost/336626.0)

# Create a new Excel workbook and add a new worksheet
wb = Workbook()
ws = wb.active

# Convert DataFrame to rows
rows = dataframe_to_rows(df, index=False, header=True)

# Write rows to the worksheet
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)

        # Apply bold font to the first row (header)
        if r_idx == 1:
            cell.font = Font(bold=True)

        # Auto-adjust column widths
        column_letter = openpyxl.utils.get_column_letter(c_idx)
        column_dimensions = ws.column_dimensions[column_letter]
        if not column_dimensions.customWidth:
            ws.column_dimensions[column_letter].width = len(str(value)) + 2

# Freeze the first row (header)
ws.freeze_panes = "A2"

# Save the workbook with the desired name
wb.save("SB 9 retention allotment analysis_05.09.23.xlsx")

# $5k for 5-9 years of experience, $10k for 10+ years of experience
df_tapr = pd.read_csv("DSTAF.csv")

# Path to the excel file
file_path = "2021-2022 TAPR DStaff Legend.xlsx"

# Use pandas to read the excel file
df_tapr_legend = pd.read_excel(file_path, engine='openpyxl')

# Create a dictionary from the dataframe
# We use strip to remove leading/trailing whitespace
data_dict = {row['NAME'].strip(): row['LABEL'].strip() for index, row in df_tapr_legend.iterrows()}

df_tapr = df_tapr.rename(columns=data_dict)

df_merged = pd.merge(df, df_tapr, on="District Number", how="left")

print(" ")
print(df_merged.head())

print(" ")

print(df_merged.columns)

